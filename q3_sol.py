# Q3_fy1_m1_three_bombs_to_excel_with_logs.py
# 任务：FY1 投 3 枚烟幕弹（间隔≥1s），最大化“遮蔽时间并集”（严格判据）。
# 方法：先在（速度、航向）小网格中挑一组，然后对三枚弹“逐枚贪心”选择（drop, fuse）。
# 日志：每步打印进度与当前并集提升；写入 result1.xlsx。

import math
import numpy as np
import openpyxl

# ---------------------- 常量与数据 ----------------------
g = 9.8
MISSILE_SPEED = 300.0
CLOUD_R = 10.0
CLOUD_SINK = 3.0
CLOUD_EFFECT = 20.0
MIN_GAP = 1.0         # 同机投弹最小间隔

CYL_BASE_CENTER = np.array([0.0, 200.0, 0.0])
CYL_R = 7.0
CYL_H = 10.0

M1_INIT = np.array([20000.0, 0.0, 2000.0])
FY1_INIT = np.array([17800.0, 0.0, 1800.0])
FAKE_TARGET_ORIGIN = np.array([0.0, 0.0, 0.0])

DT = 0.05
NPHI = 60
NZ_SIDE = 3

BASE_HEADING = math.atan2(0.0 - FY1_INIT[1], 0.0 - FY1_INIT[0])
HEADING_LIST = [BASE_HEADING + math.radians(d) for d in range(-15, 16, 3)]
SPEED_LIST = [70.0, 100.0, 140.0]
DROP_LIST = [0.0, 1.0, 2.0, 3.0, 4.0, 5.0]
FUSE_LIST = [1.0, 2.0, 3.0, 4.0, 5.0]

# ---------------------- 工具 ----------------------
def unit(v):
    n = np.linalg.norm(v)
    if n < 1e-12: return v
    return v / n

def missile_state(t, M_init):
    vdir = unit(FAKE_TARGET_ORIGIN - M_init)
    v = MISSILE_SPEED * vdir
    return M_init + v * t, v

def uav_pos_vel_at(t, uav_init, speed, heading):
    vx = speed * math.cos(heading)
    vy = speed * math.sin(heading)
    return np.array([uav_init[0] + vx*t, uav_init[1] + vy*t, uav_init[2]]), np.array([vx, vy, 0.0])

def cylinder_inside_infinite_cone(M, C, r_cloud,
                                  cyl_base_center=CYL_BASE_CENTER, R=CYL_R, H=CYL_H,
                                  n_phi=NPHI, n_z_side=NZ_SIDE):
    v = C - M
    L = np.linalg.norm(v)
    if L <= 1e-9 or r_cloud >= L: return True
    cos_alpha = math.sqrt(max(0.0, 1.0 - (r_cloud/L)**2))
    B = np.array(cyl_base_center, dtype=float)
    T = B + np.array([0.0, 0.0, H], dtype=float)
    phis = np.linspace(0.0, 2.0*math.pi, n_phi, endpoint=False)
    def ok(X):
        w = X - M
        wn = np.linalg.norm(w)
        if wn < 1e-12: return True
        cos_theta = np.dot(w, v) / (wn * L)
        return cos_theta + 1e-12 >= cos_alpha
    for phi in phis:
        if not ok(B + np.array([R*math.cos(phi), R*math.sin(phi), 0.0])): return False
        if not ok(T + np.array([R*math.cos(phi), R*math.sin(phi), 0.0])): return False
    if n_z_side >= 2:
        for k in range(n_z_side):
            z = H * (k/(n_z_side - 1.0))
            center = B + np.array([0.0, 0.0, z])
            for phi in phis:
                if not ok(center + np.array([R*math.cos(phi), R*math.sin(phi), 0.0])): return False
    return True

def explosion_from(speed, heading, drop_t, fuse_dt):
    drop_pos, uav_v = uav_pos_vel_at(drop_t, FY1_INIT, speed, heading)
    expl_xy = drop_pos[:2] + uav_v[:2] * fuse_dt
    expl_z = drop_pos[2] - 0.5 * g * (fuse_dt**2)
    return drop_pos, np.array([expl_xy[0], expl_xy[1], expl_z]), (drop_t + fuse_dt)

def sim_interval_indices(expl_t, expl_pos):
    hit_time = np.linalg.norm(M1_INIT - FAKE_TARGET_ORIGIN) / MISSILE_SPEED
    t0 = expl_t
    t1 = min(expl_t + CLOUD_EFFECT, hit_time)
    idxs = set()
    steps = int((t1 - t0) / DT) + 1
    for i in range(steps):
        t = t0 + i * DT
        M, _ = missile_state(t, M1_INIT)
        C = np.array([expl_pos[0], expl_pos[1], expl_pos[2] - 3.0 * max(0.0, t - expl_t)])
        if cylinder_inside_infinite_cone(M, C, CLOUD_R):
            idxs.add(int(round(t / DT)))
    return idxs

def pick_best_heading_speed():
    print("[Q3] 选择 (heading, speed) ...")
    best = (-1.0, None)
    total = len(HEADING_LIST) * len(SPEED_LIST)
    k = 0
    # 单弹效果预评估（用最简单 drop/fuse 组合），仅用于挑选航向/速度
    for h in HEADING_LIST:
        for s in SPEED_LIST:
            k += 1
            if k % max(1, total // 10) == 0:
                print(f"    [Q3] 预评估进度 {int(100*k/total)}%")
            dp, ep, et = explosion_from(s, h, 0.0, 2.0)
            if ep[2] <= 0: continue
            cover = len(sim_interval_indices(et, ep)) * DT
            if cover > best[0]:
                best = (cover, (h, s))
                print(f"      [Q3] 暂优航向={math.degrees(h):.1f}°, 速度={s}, 单弹≈{cover:.3f}s")
    print(f"[Q3] 航向/速度已选：heading={math.degrees(best[1][0]):.2f}°, speed={best[1][1]:.1f}")
    return best[1]

def greedy_three_bombs(heading, speed):
    print("[Q3] 贪心选择三枚烟幕弹（保证间隔≥1s）...")
    chosen = []
    union_idx = set()
    for b in range(1, 4):
        best = (-1.0, None)
        total = len(DROP_LIST) * len(FUSE_LIST)
        cnt = 0
        for d in DROP_LIST:
            # 间隔约束
            ok_gap = True
            for (d_prev, _, _, _, _, _) in chosen:
                if abs(d - d_prev) < MIN_GAP:
                    ok_gap = False
                    break
            if not ok_gap: 
                continue
            for f in FUSE_LIST:
                cnt += 1
                if cnt % max(1, total // 10) == 0:
                    print(f"    [Q3] 选择第{b}枚：进度 {int(100*cnt/total)}%")
                dp, ep, et = explosion_from(speed, heading, d, f)
                if ep[2] <= 0: 
                    continue
                idxs = sim_interval_indices(et, ep)
                gain = len(union_idx | idxs) - len(union_idx)
                gain_time = gain * DT
                if gain_time > best[0]:
                    best = (gain_time, (d, f, dp, ep, et, idxs))
        if best[1] is None:
            print(f"    [Q3] 第{b}枚未找到有效提升，提前结束。")
            break
        gain_time, (d, f, dp, ep, et, idxs) = best
        union_idx |= idxs
        chosen.append((d, f, dp, ep, et, idxs))
        print(f"    [Q3] 选中第{b}枚：drop={d:.2f}s, fuse={f:.2f}s, 起爆@{ep}, 增益≈{gain_time:.3f}s")
    total_time = len(union_idx) * DT
    print(f"[Q3] 三枚弹遮蔽并集≈{total_time:.3f}s")
    # 整理由于 Excel 输出
    out = []
    for i, (d, f, dp, ep, et, idxs) in enumerate(chosen, start=1):
        out.append({
            "num": i, "drop_t": d, "fuse": f, "drop_pos": dp, "expl_pos": ep, "expl_t": et,
            "cover": len(idxs) * DT,
            "heading_deg": math.degrees(heading), "speed": speed
        })
    return total_time, out

def write_result1_xlsx(records, path="result1.xlsx"):
    print(f"[Q3] 写入 Excel：{path}")
    try:
        wb = openpyxl.load_workbook(path)
        ws = wb.active
    except Exception:
        wb = openpyxl.Workbook()
        ws = wb.active
    ws.delete_rows(1, ws.max_row)
    ws.append(["Heading(deg)","Speed(m/s)","Bomb#","DropX","DropY","DropZ","ExplX","ExplY","ExplZ","CoverTime(s)","DropTime(s)","Fuse(s)","ExplTime(s)"])
    for r in records:
        dp, ep = r["drop_pos"], r["expl_pos"]
        ws.append([
            float(r["heading_deg"]),
            float(r["speed"]),
            int(r["num"]),
            float(dp[0]), float(dp[1]), float(dp[2]),
            float(ep[0]), float(ep[1]), float(ep[2]),
            float(r["cover"]),
            float(r["drop_t"]), float(r["fuse"]), float(r["expl_t"])
        ])
    wb.save(path)
    print("[Q3] Excel 写入完成。")

if __name__ == "__main__":
    print("[Q3] 开始")
    h, s = pick_best_heading_speed()
    total, recs = greedy_three_bombs(h, s)
    for r in recs:
        print(f"  [Q3] 弹{r['num']}: drop={r['drop_t']:.2f}s fuse={r['fuse']:.2f}s cover≈{r['cover']:.3f}s")
    write_result1_xlsx(recs, "result1.xlsx")
    print("[Q3] 结束")
